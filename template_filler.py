"""
Template Filler: opens the real OI budget template and fills only the
input cells (col C = personnel, col E = units/days, col J = calculator).
All formulas in col G stay untouched.
"""
import math
import shutil
import openpyxl
import openpyxl.cell.cell as _opc


# ── Fixed rules learned from actual budgets ───────────────────────────
SURVEYS_PER_DAY       = 10     # standard; lower for long surveys
PAGES_PER_TOOL        = 25     # always 25 regardless of duration
TRAINING_MATERIAL_PGS = 25     # always 25
MONITORING_FOOD_DAYS  = 3      # researcher monitoring visit
MONITORING_ACC_NIGHTS = 2      # researcher monitoring visit
FIELD_COORD_MONTHS    = 1      # fixed
TRANSCRIPTION_UNITS   = 1      # fixed cost, not per interview
DATA_OPS_MONTHS       = 1      # fixed
OVERHEAD_PCT          = 0.03   # 3% (confirmed from BITS/Adani)


def fill_template(schema: dict, template_path: str, output_path: str):
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Budget Internal "]

    # ── Derived values ─────────────────────────────────────────────
    sample      = schema["sample_size"]
    n_states    = max(len(schema["states"]), 1)
    n_blocks    = max(schema.get("num_blocks", 1), 1)
    dur         = schema["survey_duration"]
    spd         = _surveys_per_day(dur)
    sample_per_state = max(schema["sample_size"] // n_states, 1)
    if n_blocks > 1:
        # Block-based teams (e.g. BITS: 9 blocks → 18 enumerators, 9 supervisors)
        n_enum = n_blocks * 2
        n_sup  = n_blocks
    else:
        # Sample-based: aim to finish each state in ~2 field days
        target_days = 2
        n_enum = max(math.ceil(sample_per_state / (spd * target_days)), 2)
        n_sup  = max(math.ceil(n_enum / 5), 1)
    n_fgd       = schema.get("num_fgds", 0)
    n_idi       = schema.get("num_idis", 0)
    n_qual      = n_fgd + n_idi
    months      = schema.get("timeline_months", 3)
    components  = schema.get("components", ["Data Collection"])
    oi_codes    = schema.get("oi_codes", True)
    oi_devices  = schema.get("oi_devices", True)

    # Translation pages: 25 per tool × n_languages + 25 training
    n_tools = 1 + (1 if n_fgd > 0 else 0) + (1 if n_idi > 0 else 0)
    n_lang  = max(len(schema.get("languages", [])), 1)
    # 100 pages per language (covers all tools + training material)
    trans_pages = 100 * n_lang

    # Researcher monitoring visit days (scale slightly with states)
    mon_food = MONITORING_FOOD_DAYS + max(n_states - 1, 0)
    mon_acc  = MONITORING_ACC_NIGHTS + max(n_states - 1, 0)
    mon_cabs = MONITORING_FOOD_DAYS + max(n_states - 1, 0)

    # Accounting months scales with states
    acc_months = 1 if n_states == 1 else 2

    # ── Build label map ─────────────────────────────────────────────
    lmap = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        b = row[1].value
        if b and isinstance(b, str):
            key = b.strip().lower()
            lmap.setdefault(key, []).append(row[0].row)

    # ── 1. STUDY PREPARATION ───────────────────────────────────────
    _set(ws, lmap, "translation",          e=trans_pages,              after=1)
    _set(ws, lmap, "project manager",      c=1, e=4,                   after=1)
    _set(ws, lmap, "senior researcher",    c=1, e=5,                   after=1)
    _set(ws, lmap, "junior researcher",    c=1, e=8,                   after=1)
    _set(ws, lmap, "coder",                c=1 if oi_codes else 0,
                                           e=2 if oi_codes else 0,     after=1)

    # ── 2. J-COLUMN FIELD TEAM CALCULATOR ─────────────────────────
    j_base = _find_row(ws, "field team", after=60)
    if j_base:
        _jset(ws, j_base,    sample)      # total sample
        _jset(ws, j_base+1,  n_blocks)    # teams/blocks
        _jset(ws, j_base+2,  n_enum)      # enumerators
        _jset(ws, j_base+3,  n_sup)       # supervisors
        _jset(ws, j_base+4,  spd)         # surveys per day
        # j_base+5 is formula for DC days — skip
        _jset(ws, j_base+6,  3)           # buffer days
        _jset(ws, j_base+7,  2)           # training days
        if n_qual > 0:
            _jset(ws, j_base+11, n_qual)  # total qual interviews
            _jset(ws, j_base+12, 1)       # qual field workers
            _jset(ws, j_base+13, 2)       # qual interviews/day
            _jset(ws, j_base+14, 2)       # qual training days

    # ── 3. RESEARCH PERSONNEL (monitoring, in 2.2.2) ──────────────
    rp = _find_row(ws, "research personnel", after=60)
    if rp:
        sr = _find_after(ws, "senior researcher", rp)
        jr = _find_after(ws, "junior researcher",  rp)
        fc = _find_after(ws, "field coordination", rp)
        if sr: _write(ws, sr, c=1, e=4)                    # SR: 4 days monitoring
        if jr: _write(ws, jr, c=1, e=8)                    # JR: 8 days monitoring
        if fc: _write(ws, fc, e=FIELD_COORD_MONTHS)        # fixed 1 month

    # ── 4. TRAINING (State-1) ──────────────────────────────────────
    tr = _find_row(ws, "field training", after=60)
    if tr:
        th  = _find_after(ws, "training hall", tr)
        dev = _find_after(ws, "devices",        tr)
        if th:  _write(ws, th,  e=1)
        if dev: _write(ws, dev, e=1)

    # ── 5. LOGISTICS (State-1) ────────────────────────────────────
    lg = _find_row(ws, "logistics for researchers", after=60)
    if lg:
        flt  = _find_after(ws, "flight charges",         lg)
        cab  = _find_after(ws, "cab fare to destination", lg)
        food = _find_after(ws, "food",                    lg)
        acc  = _find_after(ws, "core team accomodation",  lg)
        lcab = _find_after(ws, "local travel cabs per sub team", lg)
        if flt:  _write(ws, flt,  c=2, e=2)                      # 2 people, 2 trips
        if cab:  _write(ws, cab,  c=2, e=2)
        if food: _write(ws, food, c=2, e=mon_food)               # monitoring visit
        if acc:  _write(ws, acc,  c=2, e=mon_acc)
        if lcab: _write(ws, lcab, c=2, e=mon_cabs)

    # ── 6. DATA MANAGEMENT ────────────────────────────────────────
    dm = _find_row(ws, "data management", after=150)
    if dm:
        trs = _find_after(ws, "transcription",            dm)
        pm  = _find_after(ws, "project manager",          dm)
        sr  = _find_after(ws, "senior researcher (quant)", dm)
        jr  = _find_after(ws, "junior researcher (quant)", dm)
        sr2 = _find_after(ws, "senior researcher (qual)",  dm)
        jr2 = _find_after(ws, "junior researcher (qual)",  dm)
        ops = _find_after(ws, "data management costs",    dm)

        # 1 unit for single-state; n_qual per state for multi-state
        trs_units = n_qual if (n_qual > 0 and n_states > 1) else (1 if n_qual > 0 else 0)
        if trs: _write(ws, trs, e=trs_units)
        if pm:  _write(ws, pm,  c=1, e=4)
        if sr:  _write(ws, sr,  c=1, e=5)
        if sr2 and n_qual > 0: _write(ws, sr2, c=1, e=2)
        if jr2 and n_qual > 0: _write(ws, jr2, c=1, e=3)
        if ops: _write(ws, ops, e=DATA_OPS_MONTHS)

    # ── 7. ANALYSIS ───────────────────────────────────────────────
    if "Analysis" in components:
        # Analysis rows are same as data mgmt rows in BITS template
        # They share the section — already handled above
        pass

    # ── 8. DEVICES ────────────────────────────────────────────────
    dv = _find_row(ws, "devices and software", after=150)
    if not dv:
        dv = _find_row(ws, "devices", after=150)
    if dv:
        cto = _find_after(ws, "surveycto",     dv)
        tab = _find_after(ws, "tablet",         dv)
        vr  = _find_after(ws, "voice recorder", dv)
        if cto: _write(ws, cto, e=1 if oi_codes else 0)
        if tab: _write(ws, tab, e=n_enum * n_states if oi_devices else 0)
        vr_units = (1 if n_states == 1 else n_states * 2) if n_qual > 0 else 0
        if vr:  _write(ws, vr,  e=vr_units)

    # ── 9. ADMIN ──────────────────────────────────────────────────
    adm = _find_row(ws, "administrative", after=150)
    if adm:
        cou = _find_after(ws, "courier",             adm)
        prn = _find_after(ws, "printing",            adm)
        acc = _find_after(ws, "accounting personnel", adm)
        leg = _find_after(ws, "legal",               adm)
        import math as _m
        if cou: _write(ws, cou, e=_m.ceil(n_states/2))
        if prn: _write(ws, prn, e=n_states)
        if acc: _write(ws, acc, e=acc_months)
        if leg: _write(ws, leg, e=1)

    wb.save(output_path)


# ── Helpers ────────────────────────────────────────────────────────────

def _find_row(ws, keyword, after=1):
    kw = keyword.lower()
    for row in ws.iter_rows(min_row=after, max_row=ws.max_row):
        b = row[1].value
        if b and isinstance(b, str) and kw in b.lower():
            return row[0].row
    return None


def _find_after(ws, keyword, after_row, window=80):
    kw = keyword.lower()
    for row in ws.iter_rows(min_row=after_row, max_row=min(after_row+window, ws.max_row)):
        b = row[1].value
        if b and isinstance(b, str) and kw in b.lower():
            return row[0].row
    return None


def _set(ws, lmap, label, c=None, e=None, after=1):
    rows = lmap.get(label.lower(), [])
    for r in rows:
        if r >= after:
            _write(ws, r, c=c, e=e)
            return


def _write(ws, row, c=None, e=None):
    if c is not None:
        cell = ws.cell(row, 3)
        if not isinstance(cell, _opc.MergedCell):
            cell.value = c
    if e is not None:
        cell = ws.cell(row, 5)
        if not isinstance(cell, _opc.MergedCell):
            cell.value = e


def _jset(ws, row, val):
    cell = ws.cell(row, 10)
    if not isinstance(cell, _opc.MergedCell):
        cell.value = val


def _surveys_per_day(duration_mins):
    # OI standard: 10 surveys/enumerator/day regardless of duration
    # This accounts for travel, refusals, and supervisor overhead
    return 10
